"""COB 30-Jun-2026 report: Daily PnL + MTD by account + EOD positions.

Sources: pnl_cob0630.txt (daily ALL-TRADES table), mtd_cob0630.txt
(mtd_split --by-account), trades_spot_avgcost (avg cost as of the cutoff).
Output: ptf8041_report_COB30Jun2026.xlsx — new workbook, openpyxl is safe.
"""
from __future__ import annotations

from decimal import Decimal as D

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

import avgcost_db

GREEN, RED = "FF107C41", "FFC00000"
CUTOFF = "2026-06-30 23:59:59.999+00"


def box_rows(path):
    rows = []
    for ln in open(path, encoding="utf-8"):
        if not ln.startswith("│"):
            continue
        cells = [c.strip() for c in ln.strip("│\n").split("│")]
        rows.append(cells)
    return rows


def num(s):
    s = s.replace(",", "").replace("—", "")
    try:
        return float(s)
    except ValueError:
        return None


def style_sheet(ws, header_row=1, pnl_cols=(), qty_cols=(), widths=None):
    for c in ws[header_row]:
        c.font = Font(bold=True)
        c.alignment = Alignment(horizontal="center")
    for r in ws.iter_rows(min_row=header_row + 1):
        for c in r:
            if not isinstance(c.value, (int, float)):
                continue
            c.number_format = ("#,##0.0000" if c.column in qty_cols
                               else "#,##0.00")
            if c.column in pnl_cols and c.value:
                c.font = Font(color=GREEN if c.value > 0 else RED)
    for i, col in enumerate(ws.columns, 1):
        want = (widths or {}).get(i)
        if want is None:
            want = min(34, max(len(str(c.value or "")) for c in col) + 2)
        ws.column_dimensions[get_column_letter(i)].width = want


def sheet_daily(wb):
    ws = wb.create_sheet("Daily PnL 30 Jun")
    ws.append(["PORTFOLIO 8041 — DAILY PnL (ALL TRADES) — "
               "COB 2026-06-30 23:59:59 UTC"])
    ws["A1"].font = Font(bold=True, size=12)
    ws.append([])
    hdr = ["venue", "account", "instrument", "SOD qty", "EOD qty", "EOD Px",
           "realized", "dUnreal", "funding", "fees", "net PnL"]
    ws.append(hdr)
    rows = box_rows("pnl_cob0630.txt")
    for c in rows:
        if len(c) < 13 or c[0] in ("venue", ""):
            continue
        vals = [num(x) for x in (c[5], c[6], c[7], c[8], c[9], c[10],
                                 c[11], c[12])]
        if c[0] == "TOTAL":
            ws.append(["TOTAL", "", "", None, None, None,
                       vals[3], vals[4], vals[5], vals[6], vals[7]])
            for cell in ws[ws.max_row]:
                cell.font = Font(bold=True)
            continue
        if all(not v for v in (vals[3], vals[4], vals[5], vals[6], vals[7])) \
                and not vals[0] and not vals[1]:
            continue                       # dead line: no qty, no pnl
        ws.append([c[0], c[1], c[2]] + vals)
    style_sheet(ws, header_row=3, pnl_cols=(7, 8, 11), qty_cols=(4, 5))
    ws.freeze_panes = "A4"
    return ws


def sheet_mtd(wb):
    ws = wb.create_sheet("MTD by Account")
    ws.append(["PORTFOLIO 8041 — MTD PnL BY ACCOUNT — 2026-06-12 (inception) "
               "-> COB 2026-06-30   (June MTD = ITD)"])
    ws["A1"].font = Font(bold=True, size=12)
    ws.append([])
    rows = box_rows("mtd_cob0630.txt")
    hdr = rows[0]
    ws.append(hdr)
    for c in rows[1:]:
        if c == hdr:
            continue
        vals = [c[0]] + [num(x) for x in c[1:]]
        ws.append(vals)
        if c[0].startswith("ITD"):
            for cell in ws[ws.max_row]:
                cell.font = Font(bold=True)
    ncols = len(hdr)
    style_sheet(ws, header_row=3, pnl_cols=tuple(range(2, ncols + 1)))
    ws.freeze_panes = "B4"
    return ws


def sheet_positions(wb):
    ws = wb.create_sheet("EOD Positions 30 Jun")
    ws.append(["PORTFOLIO 8041 — EOD BOOK POSITIONS — as of "
               "2026-06-30 23:59:59 UTC (avg-cost store)"])
    ws["A1"].font = Font(bold=True, size=12)
    ws.append([])
    ws.append(["venue", "account", "instrument", "position qty", "avg cost",
               "EOD Px", "MTM value", "unreal PnL"])
    # EOD marks by (account, display instrument) from the daily table
    marks = {}
    for c in box_rows("pnl_cob0630.txt"):
        if len(c) >= 13 and c[0] not in ("venue", "TOTAL", ""):
            marks[(c[1], c[2])] = num(c[7])
    conn = avgcost_db.connect()
    try:
        with conn.cursor() as cur:
            cur.execute("""
                SELECT DISTINCT ON (instrument, account)
                       venue, account, instrument, pos_qty_after,
                       avg_cost_after
                FROM trades_spot_avgcost
                WHERE trade_date <= %s
                ORDER BY instrument, account, trade_date DESC, id DESC
            """, (CUTOFF,))
            recs = cur.fetchall()
    finally:
        conn.close()
    out = []
    for venue, acct, inst, qty, avg in recs:
        if abs(qty) < D("0.000001"):
            continue
        disp = inst.split("@")[0]
        # match the daily table's display naming to borrow its EOD mark
        cands = (disp, disp.split("/")[0], disp.replace("-P/", ""),
                 disp.replace("-P/", "").replace("/", ""))
        mark = next((marks[(acct, k)] for k in cands
                     if (acct, k) in marks), None)
        q, a = float(qty), float(avg)
        mtm = q * mark if mark else None
        upnl = (mark - a) * q if mark else None
        out.append([venue, acct, disp, q, a, mark, mtm, upnl])
    out.sort(key=lambda r: -abs(r[6] or 0))
    for r in out:
        ws.append(r)
    tot_mtm = sum(r[6] for r in out if r[6])
    tot_u = sum(r[7] for r in out if r[7])
    ws.append(["TOTAL (marked rows)", "", "", None, None, None, tot_mtm,
               tot_u])
    for cell in ws[ws.max_row]:
        cell.font = Font(bold=True)
    style_sheet(ws, header_row=3, pnl_cols=(8,), qty_cols=(4, 5, 6))
    ws.freeze_panes = "A4"
    return ws


def main():
    wb = Workbook()
    wb.remove(wb.active)
    sheet_daily(wb)
    sheet_mtd(wb)
    sheet_positions(wb)
    out = "ptf8041_report_COB30Jun2026.xlsx"
    wb.save(out)
    print("wrote", out)
    for ws in wb.worksheets:
        print(f"  {ws.title}: {ws.max_row - 3} data rows")


if __name__ == "__main__":
    main()
