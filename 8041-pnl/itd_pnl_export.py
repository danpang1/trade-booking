"""Export ITD PnL by venue/asset + the trades behind it to Excel.

  python itd_pnl_export.py --date 2026-07-01

Sheet 1 "ITD PnL"  — one row per (venue, account, asset) leg over ALL stored
  trades up to the COB EOD boundary: trades, EOD qty, avg cost, EOD price used
  (COB marks: HL xyz index / Yahoo EOD / Binance snap index; stables @ 1),
  realized, unrealized, funding, fees, ITD net. IBKR manual overlay appended.
Sheet 2 "Trades"   — every stored fill <= COB EOD with its price, fees,
  engine realized and running position/avg-cost.
"""
from __future__ import annotations
import argparse
import sys
from datetime import datetime, timedelta, timezone
from decimal import Decimal
from pathlib import Path

try:
    sys.stdout.reconfigure(encoding="utf-8")
except Exception:
    pass

REPO = Path(__file__).resolve().parent
sys.path.insert(0, str(REPO))
import avgcost_db as adb  # noqa: E402
import equity_marks as eqm  # noqa: E402
import account_recon as ar  # noqa: E402
from engine import Position, unrealized, D, ZERO  # noqa: E402
from native_common import ACCOUNT_NAME as NTV_ACCT, SYMBOL_TO_HL  # noqa: E402
from asset_map import std_asset  # noqa: E402

BIN_ACCT = "TK810@BINANCE_USDT_FUTURE"
HLF_ACCT = "TRADING_06@HYPERLIQUID_FUTURES"
PINNED_MARKS = {"2026-06-14": "168.01", "2026-06-15": "199.82"}
STABLES = {"USDT", "USDC", "USD", "USDG"}
# Manually supplied venue day figures (hedge account, no trades in the store).
IBKR_DAYS = {"2026-06-30": Decimal("-20473.5"), "2026-07-01": Decimal("56628.31"),
             "2026-07-02": Decimal("58476"),
             # 07-03..05 combined lump: user ITD@07-05 56,299.00 − 94,630.81
             "2026-07-05": Decimal("-38331.81"),
             # user ITD@07-06 64,919.00 − 56,299.00
             "2026-07-06": Decimal("8620"),
             # user ITD@07-07 143,004 (balance 693,004 − 550,000 initial funding)
             "2026-07-07": Decimal("78085"),
             # user ITD@07-08 114,228
             "2026-07-08": Decimal("-28776"),
             # user ITD@07-09 51,322
             "2026-07-09": Decimal("-62906"),
             # 07-10..12 combined lump: user ITD@07-12 32,881 − 51,322
             "2026-07-12": Decimal("-18441"),
             # user ITD@07-13 128,712
             "2026-07-13": Decimal("95831"),
             # user ITD@07-14 57,967.78
             "2026-07-14": Decimal("-70744.22"),
             # user ITD@07-15 86,701.42
             "2026-07-15": Decimal("28733.64")}

ap = argparse.ArgumentParser(description="8041 ITD PnL by venue/asset -> xlsx")
ap.add_argument("--date", default="2026-07-01", help="COB day YYYY-MM-DD")
ap.add_argument("--inception", default="2026-06-12",
                help="ITD window start (book inception; realized/funding/fees "
                     "before this are excluded — position/avg-cost still carry in)")
ap.add_argument("--out", default=None, help="output .xlsx path")
ap.add_argument("--pins", default=None,
                help="CSV (symbol,price) of official EOD marks — takes "
                     "precedence over every feed. Per leg the plain ticker is "
                     "tried first, then the 'X' variant (xyz-style marks); "
                     "Native {SYM}B and SPCXD map to the underlying ticker.")
ap.add_argument("--prev-pins", default=None,
                help="pins CSV for the PREVIOUS COB (Day PnL column marks); "
                     "defaults to eod_pins_<prev-date>.csv when it exists")
ap.add_argument("--prev-cob", default=None, metavar="YYYY-MM-DD",
                help="override the Day-PnL start boundary: 'Day' figures "
                     "become the cumulative window (prev-cob, COB] instead of "
                     "the single COB day (e.g. --date 2026-07-12 "
                     "--prev-cob 2026-07-09 -> 3-day PnL 10-12 Jul)")
a = ap.parse_args()
COB = a.date
OUT = Path(a.out) if a.out else REPO / f"8041_itd_pnl_cob{COB}.xlsx"

w0 = datetime.strptime(a.inception, "%Y-%m-%d").replace(tzinfo=timezone.utc)
w1 = (datetime.strptime(COB, "%Y-%m-%d").replace(tzinfo=timezone.utc)
      + timedelta(days=1))
if a.prev_cob:
    PREV = a.prev_cob                                    # window start COB day
    w1p = (datetime.strptime(PREV, "%Y-%m-%d").replace(tzinfo=timezone.utc)
           + timedelta(days=1))                          # its EOD boundary
    WINLBL = f"{PREV[5:]}->{COB[5:]} {COB[:4]}"
else:
    w1p = w1 - timedelta(days=1)                         # previous COB EOD
    PREV = (w1p - timedelta(days=1)).strftime("%Y-%m-%d")  # previous COB day
    WINLBL = COB
BOUNDARY = w1.replace(tzinfo=None)
BOUNDARY_P = w1p.replace(tzinfo=None)
if a.prev_pins is None:
    _pp = REPO / f"eod_pins_{PREV}.csv"
    a.prev_pins = str(_pp) if _pp.exists() else None


def hist_index(inst_like, account=HLF_ACCT, boundary=None):
    """Venue snapshot index_price at an EOD boundary (tq_hist_position)."""
    import psycopg2
    from pg import PG_HOST, PG_PORT, PG_USER, PG_PASS, PG_DB
    boundary = boundary if boundary is not None else BOUNDARY
    conn = psycopg2.connect(host=PG_HOST, port=PG_PORT, user=PG_USER,
                            password=PG_PASS, database=PG_DB, connect_timeout=15)
    try:
        cur = conn.cursor()
        cur.execute("""
            SELECT index_price FROM tq_hist_position
            WHERE account_name = %s AND instrument ILIKE %s
              AND record_ts >= %s AND record_ts < %s
            ORDER BY record_ts LIMIT 1
        """, (account, inst_like, boundary, boundary + timedelta(minutes=60)))
        r = cur.fetchone()
        return D(str(r[0])) if r and r[0] is not None else None
    finally:
        conn.close()


def load_pins(path):
    import csv as _csv
    pins = {}
    with open(path, encoding="utf-8-sig", newline="") as fh:
        for r in _csv.DictReader(fh):
            pins[r["symbol"].strip().upper()] = D(r["price"].strip())
    print(f"loaded {len(pins)} pinned EOD marks from {path}")
    return pins


PINS = load_pins(a.pins) if a.pins else {}
PINS_P = load_pins(a.prev_pins) if a.prev_pins else {}


def _pin_for(inst, pins):
    """Pinned mark for a leg: plain ticker first, then the 'X' xyz variant."""
    base = inst.split("/")[0]
    if base.startswith("xyz:"):
        t = base[4:].split("-P")[0]
        cands = (t + "X", t)                 # xyz legs prefer the X variant
    elif "@NATIVECORE" in inst:
        hl = SYMBOL_TO_HL.get(base)
        t = hl.split("-P/")[0].replace("xyz:", "") if hl else base
        cands = (t, t + "X")
    elif base == "SPCXD":
        cands = ("SPCX",)
    elif "@BINANCE_SPOT" in inst or "@ETHEREUM_RFQ" in inst:
        t = std_asset(base)                  # SPCXB -> SPCX, MUB -> MU, ETH -> ETH
        cands = (t, t + "X")
    else:
        t = base.split("-P")[0]
        cands = (t, t + "X")
    for c in cands:
        if c in pins:
            return pins[c]
    return None


def eod_mark(inst, *, boundary=None, date_iso=None, pins=None, is_cob=True):
    """(mark, source) for a leg at an EOD boundary — user pins first, then the
    same routing as the daily table."""
    boundary = boundary if boundary is not None else BOUNDARY
    date_iso = date_iso or COB
    pins = PINS if pins is None else pins
    base = inst.split("/")[0].split("-P")[0]
    if pins:
        p = _pin_for(inst, pins)
        if p is not None:
            return p, "user pin"
    if base in STABLES:
        return D(1), "stable@1"
    if "@NATIVECORE" in inst:
        hl = SYMBOL_TO_HL.get(inst.split("/")[0])
        if hl:
            m = hist_index(hl.split("-P/")[0] + "%", boundary=boundary)
            return m, "HL xyz index"
        return None, "none"
    if "@BITSTAMP" in inst or "@ROBINHOOD" in inst:
        t = inst.split("/")[0]
        m = eqm.resolve_mark(t, date_iso, is_cob=is_cob,
                             hist_fn=lambda tk, d: hist_index("xyz:" + tk + "%",
                                                              boundary=boundary))
        return m, ("Yahoo/xyz EOD" if m is not None else "none")
    if inst.startswith("SPCX-P"):                        # Binance SPCX perp
        m = (hist_index("%SPCX%", BIN_ACCT, boundary)
             or (D(PINNED_MARKS[date_iso]) if date_iso in PINNED_MARKS else None))
        return m, "Binance snap index"
    if "SPCX" in inst:                                   # HL SPCXD spot pair leg —
        m = hist_index("%SPCX%", BIN_ACCT, boundary)     # same mark as the perp side
        return m, "Binance snap index"
    if "@BINANCE_SPOT" in inst:                          # B-token tokenized equity
        t = std_asset(inst.split("/")[0])
        m = eqm.resolve_mark(t, date_iso, is_cob=is_cob,
                             hist_fn=lambda tk, d: hist_index("xyz:" + tk + "%",
                                                              boundary=boundary))
        return m, ("Yahoo/xyz EOD" if m is not None else "none")
    if "@ETHEREUM_RFQ" in inst:                          # ETH inventory leg
        m = hist_index("ETH%", BIN_ACCT, boundary)       # tk810 ETH perp index
        return m, "Binance snap index"
    if "@BINANCE" in inst:                               # non-SPCX Binance UM perp
        m = hist_index(inst.split("-P/")[0] + "%", BIN_ACCT, boundary)
        return m, "Binance snap index"
    m = hist_index(inst.split("-P/")[0] + "%", boundary=boundary)  # HL perp legs
    return m, "HL snap index"


# ── funding (raw rows; summed per window per leg) ───────────────────
print("pulling funding (Binance income + HL ledger)...")
_bin_income = [r for r in ar.bin_income_raw(int(w0.timestamp() * 1000),
                                            int(w1.timestamp() * 1000))
               if r.get("incomeType") == "FUNDING_FEE"]
_hl_funding = ar.hl_funding()


def bin_fund_sym(symbol, t1):
    return sum((D(str(r["income"])) for r in _bin_income
                if str(r.get("symbol", "")) == symbol
                and int(r["time"]) < t1.timestamp() * 1000), ZERO)


def hl_fund_coin(coin, t1):
    return sum((D(str(r["delta"]["usdc"])) for r in _hl_funding
                if r["delta"]["coin"] == coin
                and w0.timestamp() * 1000 <= int(r["time"]) < t1.timestamp() * 1000), ZERO)


def leg_funding(inst, prod, acct, t1):
    if "@BINANCE" in inst and "-P/" in inst:     # UM perps only (spot: no funding)
        return bin_fund_sym(inst.split("-P/")[0]
                            + inst.split("-P/")[1].split("@")[0], t1)
    if prod == "PERP" and acct == HLF_ACCT:
        return hl_fund_coin(inst.split("-P/")[0], t1)
    return ZERO

# ── per-leg ITD ─────────────────────────────────────────────────────
print("computing per-leg ITD...")
conn = adb.connect()
rows = []
try:
    # group venues together: venue -> account -> instrument
    for inst, acct, prod, quote in sorted(
            adb.distinct_instruments(conn),
            key=lambda r: (ar.venue_of(r[1]), r[1], r[0])):
        realized, fee_usd, na, nm = adb.window_agg(conn, inst, w0, w1)
        eq, ea = adb.pos_at(conn, inst, w1)
        eq, ea = D(str(eq)), D(str(ea))
        mark, msrc = eod_mark(inst)
        unrl = unrealized(Position(eq, ea), mark) if (mark is not None and eq != 0) else ZERO
        fnd = leg_funding(inst, prod, acct, w1)
        net = D(str(realized)) + unrl + fnd - D(str(fee_usd))
        # previous COB EOD -> Day PnL = ITD(COB) − ITD(prev)
        realized_p, fee_usd_p, na_p, nm_p = adb.window_agg(conn, inst, w0, w1p)
        pq, pa = adb.pos_at(conn, inst, w1p)
        pq, pa = D(str(pq)), D(str(pa))
        mark_p, _ = eod_mark(inst, boundary=BOUNDARY_P, date_iso=PREV,
                             pins=PINS_P, is_cob=False)
        unrl_p = unrealized(Position(pq, pa), mark_p) if (mark_p is not None and pq != 0) else ZERO
        fnd_p = leg_funding(inst, prod, acct, w1p)
        net_p = D(str(realized_p)) + unrl_p + fnd_p - D(str(fee_usd_p))
        rows.append({
            "venue": ar.venue_of(acct), "account": acct, "instrument": inst,
            "asset": std_asset(inst.split("/")[0]), "trades": na + nm,
            "eod_qty": eq, "avg_cost": ea, "eod_price": mark, "mark_src": msrc,
            "realized": D(str(realized)), "unreal": unrl, "funding": fnd,
            "fees": D(str(fee_usd)), "itd_net": net, "day_net": net - net_p,
            # COB-day window components (ITD@COB − ITD@prev) for the Day sheet
            "day_trades": (na + nm) - (na_p + nm_p),
            "day_realized": D(str(realized)) - D(str(realized_p)),
            "day_unreal": unrl - unrl_p,
            "day_funding": fnd - fnd_p,
            "day_fees": D(str(fee_usd)) - D(str(fee_usd_p)),
        })
finally:
    conn.close()

# ── trades dump ─────────────────────────────────────────────────────
print("pulling trades...")
conn = adb.connect()
cur = conn.cursor()
cur.execute("""
    SELECT trade_date, venue, account, instrument, product, direction,
           base_amount, price, quote_asset, fee_asset, fee_amount, fee_usd,
           realized, pos_qty_after, avg_cost_after, source, counterparty,
           external_trade_id
    FROM trades_spot_avgcost
    WHERE trade_date < %s
    ORDER BY account, instrument, trade_date, external_trade_id
""", (w1,))
trades = cur.fetchall()
conn.close()
print(f"  {len(trades)} trades")

# ── write xlsx ──────────────────────────────────────────────────────
from openpyxl import Workbook  # noqa: E402
from openpyxl.styles import Font  # noqa: E402
from openpyxl.utils import get_column_letter  # noqa: E402

wb = Workbook()
bold = Font(bold=True)

ws = wb.active
ws.title = "ITD PnL"
hdr = ["Venue", "Account", "Instrument", "Asset", "Trades", "EOD Qty", "Avg Cost",
       f"EOD Price {COB}", "Mark Source", "Realized", "Unrealized", "Funding",
       "Fees", "ITD Net", f"Day PnL {WINLBL}"]
ws.append(hdr)
tot = {k: ZERO for k in ("realized", "unreal", "funding", "fees", "itd_net", "day_net")}
for r in rows:
    ws.append([r["venue"], r["account"], r["instrument"], r["asset"], r["trades"],
               float(r["eod_qty"]), float(r["avg_cost"]),
               (float(r["eod_price"]) if r["eod_price"] is not None else None),
               r["mark_src"], float(r["realized"]), float(r["unreal"]),
               float(r["funding"]), float(r["fees"]), float(r["itd_net"]),
               float(r["day_net"])])
    for k in tot:
        tot[k] += r[k]
ib_itd = sum((v for d, v in IBKR_DAYS.items() if d <= COB), ZERO)
ib_day = sum((v for d, v in IBKR_DAYS.items() if PREV < d <= COB), ZERO)
ws.append(["IBKR", "IBKR (manual)",
           "manual day figures: " + ", ".join(f"{d}/{float(v):+,.2f}"
                                              for d, v in sorted(IBKR_DAYS.items())),
           "ALL", None, None, None, None, "manual",
           None, None, None, None, float(ib_itd), float(ib_day)])
tot["itd_net"] += ib_itd
tot["day_net"] += ib_day
ws.append(["TOTAL", "", "", "", "", "", "", "", "", float(tot["realized"]),
           float(tot["unreal"]), float(tot["funding"]), float(tot["fees"]),
           float(tot["itd_net"]), float(tot["day_net"])])
for c in range(1, len(hdr) + 1):
    ws.cell(1, c).font = bold
    ws.cell(ws.max_row, c).font = bold
for i, wdt in enumerate([12, 30, 32, 8, 8, 14, 12, 14, 16, 12, 12, 10, 10, 14, 14], 1):
    ws.column_dimensions[get_column_letter(i)].width = wdt
for col in "FGHJKLMNO":
    for cell in ws[col][1:]:
        cell.number_format = "#,##0.0000" if col in "FGH" else "#,##0.00"
ws.freeze_panes = "A2"
ws.auto_filter.ref = f"A1:O{ws.max_row}"

# ── Day PnL sheet: same layout as ITD PnL, but every figure is the ──
# COB-day window (component diffs of the two EOD boundaries, so it ties
# to the Day PnL column: realized + Δunreal + funding − fees = day net).
wsd = wb.create_sheet("Day PnL")
hdrd = ["Venue", "Account", "Instrument", "Asset", f"Trades {WINLBL}",
        "EOD Qty", "Avg Cost", f"EOD Price {COB}", "Mark Source", "Realized",
        "Unrealized", "Funding", "Fees", f"Day Net {WINLBL}"]
wsd.append(hdrd)
dtot = {k: ZERO for k in ("day_realized", "day_unreal", "day_funding",
                          "day_fees", "day_net")}
for r in rows:
    wsd.append([r["venue"], r["account"], r["instrument"], r["asset"],
                r["day_trades"], float(r["eod_qty"]), float(r["avg_cost"]),
                (float(r["eod_price"]) if r["eod_price"] is not None else None),
                r["mark_src"], float(r["day_realized"]),
                float(r["day_unreal"]), float(r["day_funding"]),
                float(r["day_fees"]), float(r["day_net"])])
    for k in dtot:
        dtot[k] += r[k]
wsd.append(["IBKR", "IBKR (manual)", f"manual day figure {WINLBL}", "ALL",
            None, None, None, None, "manual", None, None, None, None,
            float(ib_day)])
dtot["day_net"] += ib_day
wsd.append(["TOTAL", "", "", "", "", "", "", "", "",
            float(dtot["day_realized"]), float(dtot["day_unreal"]),
            float(dtot["day_funding"]), float(dtot["day_fees"]),
            float(dtot["day_net"])])
for c in range(1, len(hdrd) + 1):
    wsd.cell(1, c).font = bold
    wsd.cell(wsd.max_row, c).font = bold
for i, wdt in enumerate([12, 30, 32, 8, 10, 14, 12, 14, 16, 12, 12, 10, 10,
                         14], 1):
    wsd.column_dimensions[get_column_letter(i)].width = wdt
for col in "FGHJKLMN":
    for cell in wsd[col][1:]:
        cell.number_format = "#,##0.0000" if col in "FGH" else "#,##0.00"
wsd.freeze_panes = "A2"
wsd.auto_filter.ref = f"A1:N{wsd.max_row}"

# ── Position Buildup: trades-only cost-basis positions ─────────────
# Every fill contributes +signed_qty of the base asset (net of in-kind fees)
# and −signed_qty × price of the quote asset (quote-side fees deducted).
# Futures are treated identically (long 1 SPCX @ 180 => +1 SPCX / −180 USDT).
# Funding is NOT included (trades only). Balance T = Qty T × Rate T (pins).
from collections import defaultdict as _dd  # noqa: E402

bp_pos = _dd(lambda: ZERO)     # (venue, account, instrument) -> base qty
bp_cash = _dd(lambda: ZERO)    # (venue, account, quote)      -> quote cash
for t in trades:
    ven, acct, inst = t[1], t[2], t[3]
    direction, qty, px = t[5], D(str(t[6])), D(str(t[7]))
    quote = str(t[8] or "").upper()
    fa = str(t[9] or "").upper()
    fee = D(str(t[10] or 0))
    sq = qty if direction == "LONG" else -qty
    base = inst.split("/")[0]
    if fee and fa == base.upper():          # in-kind fee -> fewer tokens settle
        sq_net = sq - fee
    else:
        sq_net = sq
    bp_pos[(ven, acct, inst)] += sq_net
    bp_cash[(ven, acct, quote)] += -(qty if direction == "LONG" else -qty) * px
    if fee and fa == quote:                 # quote-side fee -> less cash
        bp_cash[(ven, acct, quote)] -= fee

leg_rate = {r["instrument"]: r["eod_price"] for r in rows}
bp_rows = []
for (ven, acct, inst), q in bp_pos.items():
    if abs(q) < D("0.000000001"):
        continue
    raw = inst.split("/")[0]
    rate = leg_rate.get(inst)
    bp_rows.append([ven, acct, inst, std_asset(raw), q, rate,
                    (q * rate if rate is not None else None)])
for (ven, acct, quote), c in bp_cash.items():
    if abs(c) < D("0.01"):
        continue
    rate = PINS.get(quote, D(1) if quote in STABLES else None)
    bp_rows.append([ven, acct, quote, quote, c, rate,
                    (c * rate if rate is not None else None)])
# funding cash adjustments (venue income/ledger, same sources as the ITD sheet)
# so cash rows tie to the venues' actual trading+funding wallet movement.
_bin_f = sum((D(str(r["income"])) for r in _bin_income
              if int(r["time"]) < w1.timestamp() * 1000), ZERO)
_hl_f = sum((D(str(r["delta"]["usdc"])) for r in _hl_funding
             if w0.timestamp() * 1000 <= int(r["time"]) < w1.timestamp() * 1000), ZERO)
if _bin_f:
    _rt = PINS.get("USDT", D(1))
    bp_rows.append([ar.venue_of(BIN_ACCT), BIN_ACCT, "USDT (funding)", "USDT",
                    _bin_f, _rt, _bin_f * _rt])
if _hl_f:
    _rt = PINS.get("USDC", D(1))
    bp_rows.append([ar.venue_of(HLF_ACCT), HLF_ACCT, "USDC (funding)", "USDC",
                    _hl_f, _rt, _hl_f * _rt])
bp_rows.sort(key=lambda r: (r[0], r[1], r[2]))

ws3 = wb.create_sheet("Position Buildup")
hdr3 = ["Venue", "Account", "Instrument", "Common Symbol", "Qty T",
        "Rate T", "Balance T (USD)"]
ws3.append(hdr3)
bp_tot = ZERO
for r in bp_rows:
    ws3.append([r[0], r[1], r[2], r[3], float(r[4]),
                (float(r[5]) if r[5] is not None else None),
                (float(r[6]) if r[6] is not None else None)])
    if r[6] is not None:
        bp_tot += r[6]
ws3.append(["TOTAL (trades + funding MTM, excl IBKR)", "", "", "", None, None,
            float(bp_tot)])
for c in range(1, len(hdr3) + 1):
    ws3.cell(1, c).font = bold
    ws3.cell(ws3.max_row, c).font = bold
for i, wdt in enumerate([12, 30, 34, 14, 16, 14, 16], 1):
    ws3.column_dimensions[get_column_letter(i)].width = wdt
for col in "EFG":
    for cell in ws3[col][1:]:
        cell.number_format = "#,##0.0000" if col in "EF" else "#,##0.00"
ws3.freeze_panes = "A2"
ws3.auto_filter.ref = f"A1:G{ws3.max_row}"

ws2 = wb.create_sheet("Trades")
hdr2 = ["Trade Date (UTC)", "Venue", "Account", "Instrument", "Product",
        "Direction", "Qty", "Price", "Quote", "Fee Asset", "Fee", "Fee USD",
        "Realized", "Pos After", "Avg Cost After", "Source", "Counterparty",
        "External Trade Id"]
ws2.append(hdr2)
for t in trades:
    row = list(t)
    if row[0] is not None and row[0].tzinfo is not None:
        row[0] = row[0].astimezone(timezone.utc).replace(tzinfo=None)
    for i in (6, 7, 10, 11, 12, 13, 14):
        row[i] = float(row[i]) if row[i] is not None else None
    ws2.append(row)
for c in range(1, len(hdr2) + 1):
    ws2.cell(1, c).font = bold
ws2.freeze_panes = "A2"
ws2.column_dimensions["A"].width = 19
for col, wdt in (("B", 12), ("C", 28), ("D", 32), ("R", 40)):
    ws2.column_dimensions[col].width = wdt
ws2.auto_filter.ref = f"A1:R{ws2.max_row}"

wb.save(OUT)
print(f"saved {OUT}  (ITD PnL {len(rows)}+1 legs, {len(trades)} trades)")
print(f"ITD TOTAL incl manual: {float(tot['itd_net']):,.2f} USD   "
      f"Day PnL {COB}: {float(tot['day_net']):,.2f} USD")
